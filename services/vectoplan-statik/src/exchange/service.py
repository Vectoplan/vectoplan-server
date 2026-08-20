"""Structural exchange adapters.

The canonical JSON remains lossless for VECTOPLAN. SAF 2.2 is generated as an
Excel workbook with official object/list names. Importer-specific round trips
must be covered by golden files before production approval.
"""

from __future__ import annotations

import hashlib
import io
import json
import math
import uuid
from datetime import UTC, datetime
from typing import Any, Mapping

from src.materials import MaterialCatalog, default_material_catalog


def build_neutral_exchange(job: Mapping[str, Any], result: Mapping[str, Any] | None = None) -> dict[str, Any]:
    fingerprint = hashlib.sha256(json.dumps(job, sort_keys=True, ensure_ascii=False).encode("utf-8")).hexdigest()[:20]
    return {
        "contract_version": "vectoplan-structural-exchange/0.2",
        "exchange_ref": f"vpsx_{fingerprint}",
        "coordinate_system": {"handedness": "right", "vertical_axis": "Z", "length_unit": "m", "force_unit": "kN"},
        "source": {"application": "vectoplan-statik", "model_revision_ref": job.get("model_revision_ref"), "job_ref": job.get("job_ref")},
        "standards_profile": job.get("standards_profile"),
        "analysis_job": job,
        "analysis_result": result,
        "adapter_targets": {
            "SAF": {"version": "2.2.0", "status": "implemented_baseline", "round_trip_certified": False},
            "IFC": {"version": "IFC4.3.2", "status": "mapping_contract_prepared", "implemented": False, "reason": "IFC Structural Analysis excludes finite-element topology and detailed mesh stresses."},
        },
    }


class Saf22Exporter:
    saf_version = "2.2.0"

    def __init__(self, materials: MaterialCatalog | None = None) -> None:
        self.materials = materials or default_material_catalog()

    @staticmethod
    def _uuid(seed: str) -> str:
        return str(uuid.uuid5(uuid.NAMESPACE_URL, f"https://vectoplan.local/statik/{seed}"))

    @staticmethod
    def _add_table(workbook: Any, name: str, headers: list[str], rows: list[list[Any]]) -> None:
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.table import Table, TableStyleInfo

        sheet = workbook.create_sheet(name)
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="173B55")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for column in sheet.columns:
            letter = column[0].column_letter
            sheet.column_dimensions[letter].width = min(42, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
        if rows:
            table = Table(displayName=name, ref=f"A1:{sheet.cell(row=len(rows) + 1, column=len(headers)).coordinate}")
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
            sheet.add_table(table)
        sheet.freeze_panes = "A2"

    @staticmethod
    def _add_property_sheet(workbook: Any, name: str, properties: list[tuple[str, Any]]) -> None:
        from openpyxl.styles import Font, PatternFill

        sheet = workbook.create_sheet(name)
        sheet.append([name, "Value"])
        for key, value in properties:
            sheet.append([key, value])
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="173B55")
        sheet.column_dimensions["A"].width = 31
        sheet.column_dimensions["B"].width = 54

    @staticmethod
    def _material_id(job: Mapping[str, Any]) -> str:
        parameters = (job.get("design") or {}).get("parameters") or {}
        for key in ("concrete_class", "steel_grade", "timber_grade", "masonry_grade", "prestress_grade"):
            if parameters.get(key):
                return str(parameters[key])
        return "C25/30"

    @staticmethod
    def _saf_material_type(kind: str) -> str:
        return {
            "reinforced_concrete": "Concrete",
            "prestressed_concrete": "Concrete",
            "steel": "Steel",
            "timber": "Timber",
            "masonry": "Masonry",
            "aluminium": "Aluminium",
        }.get(kind, "Other")

    @staticmethod
    def _load_type(action_type: str) -> str:
        return {
            "self_weight": "Standard",
            "imposed": "Static",
            "snow": "Snow",
            "wind": "Wind",
            "traffic": "Moving",
            "prestress": "Prestress",
            "fire": "Fire",
        }.get(action_type, "Standard")

    def to_bytes(self, job: Mapping[str, Any], result: Mapping[str, Any] | None = None) -> bytes:
        try:
            from openpyxl import Workbook
        except ImportError as exc:
            raise RuntimeError("SAF export requires openpyxl") from exc

        workbook = Workbook()
        workbook.remove(workbook.active)
        now = datetime.now(UTC).strftime("%Y-%m-%d %H:%M")
        project_ref = str(job.get("project_ref", "VECTOPLAN project"))
        model_revision = str(job.get("model_revision_ref", "local"))
        self._add_property_sheet(workbook, "Project", [
            ("Name", project_ref), ("Description", "Structural analysis model exported by vectoplan-statik"), ("Project nr", project_ref),
            ("Created", now), ("Last update", now), ("Project type", "Infrastructure construction" if job.get("structure_type") == "bridge" else "Building construction"),
            ("Project kind", "New building"), ("Building type", str(job.get("structure_type", "Generic"))), ("Status", "Engineering review required"),
            ("Location", ""), ("Id", self._uuid(project_ref)),
        ])
        self._add_property_sheet(workbook, "Model", [
            ("Name", str(job.get("job_ref", "analysis-model"))), ("Description", "Analytical model - uncertified exchange baseline"),
            ("Discipline", "Load-bearing structure"), ("Level of detail", "Draft"), ("Status", "Planning stage"), ("Owner", "VECTOPLAN"),
            ("Revision number", model_revision), ("Created", now[:10]), ("Last update", now[:10]), ("Source type", "Structural analysis model"),
            ("Source application", "vectoplan-statik 0.2"), ("SAF Version", self.saf_version), ("Source company", "VECTOPLAN"),
            ("Global coordinate system", "Z vertical"), ("LCS of cross-section", "ZYX"), ("System of units", "Metric"),
            ("National code", "EC-DIN-EN (German NA)"), ("Ignored objects", ""), ("Ignored groups", ""), ("Id", self._uuid(f"model:{model_revision}")),
        ])

        material_id = self._material_id(job)
        try:
            material = self.materials.get(material_id)
        except ValueError:
            material = self.materials.get("C25/30")
        material_name = material.material_id
        self._add_table(workbook, "StructuralMaterial", ["Name", "Type", "Subtype", "Quality", "Unit mass [kg/m3]", "E modulus [MPa]", "G modulus [MPa]", "Poisson Coefficient", "Thermal expansion [1/K]", "Design properties", "Id"], [[
            material_name, self._saf_material_type(str(job.get("material_kind", ""))), "", material_name, material.density_kn_m3 * 100.0,
            material.elastic_modulus_mpa, material.elastic_modulus_mpa / (2.0 * (1.0 + material.poisson_ratio)), material.poisson_ratio,
            material.thermal_expansion_1_k, "", self._uuid(f"material:{material_name}"),
        ]])

        load_case_rows = []
        load_group_rows = []
        for case in job.get("load_cases", []):
            case_id = str(case["load_case_id"])
            group_name = f"LG_{case_id}"
            action_type = "Permanent" if case["category"] == "permanent" else "Accidental" if case["category"] == "accidental" else "Variable"
            load_group_rows.append([group_name, action_type, "Standard", "Together", self._uuid(f"load-group:{case_id}")])
            load_case_rows.append([case_id, case.get("label", case_id), action_type, group_name, self._load_type(str(case.get("action_type", ""))), "Medium" if action_type == "Variable" else "", self._uuid(f"load-case:{case_id}")])
        self._add_table(workbook, "StructuralLoadGroup", ["Name", "Action type", "Load group type", "Relation", "Id"], load_group_rows)
        self._add_table(workbook, "StructuralLoadCase", ["Name", "Description", "Action type", "Load group", "Load type", "Duration", "Id"], load_case_rows)

        model = job.get("analysis_model") or {}
        if model.get("kind") == "beam_line":
            self._add_beam_model(workbook, job, result, material_name)
        elif model.get("kind") == "surface_plate":
            self._add_surface_model(workbook, job, result, material_name)
        self._add_combinations(workbook, result)

        info = workbook.create_sheet("VECTOPLAN_Readme", 0)
        info.append(["VECTOPLAN SAF 2.2 baseline export"])
        info.append(["Status", "Engineering exchange draft - importer round-trip not certified"])
        info.append(["Source", "https://www.saf.guide/en/stable/"])
        info.append(["Important", "Check axes, releases, supports, load directions, profiles and national-code settings in the receiving application."])
        info.column_dimensions["A"].width = 22
        info.column_dimensions["B"].width = 105
        stream = io.BytesIO()
        workbook.save(stream)
        return stream.getvalue()

    def _add_beam_model(self, workbook: Any, job: Mapping[str, Any], result: Mapping[str, Any] | None, material_name: str) -> None:
        model = job["analysis_model"]
        spans = list(model.get("spans", []))
        nodes = []
        x = 0.0
        nodes.append(["N1", x, 0.0, 0.0, self._uuid("node:1")])
        for index, span in enumerate(spans, start=2):
            x += float(span["length_m"])
            nodes.append([f"N{index}", x, 0.0, 0.0, self._uuid(f"node:{index}")])
        self._add_table(workbook, "StructuralPointConnection", ["Name", "Coordinate X [m]", "Coordinate Y [m]", "Coordinate Z [m]", "Id"], nodes)

        design_parameters = (job.get("design") or {}).get("parameters") or {}
        area_m2 = float(design_parameters.get("area_mm2", design_parameters.get("width_mm", 300.0) * design_parameters.get("height_mm", 500.0))) / 1_000_000.0
        cross_rows = []
        member_rows = []
        load_rows = []
        for index, span in enumerate(spans, start=1):
            inertia = float(span["inertia_m4"])
            height = math.sqrt(max(1e-12, 12.0 * inertia / area_m2))
            width = area_m2 / height
            cross_name = f"CS{index}"
            cross_rows.append([cross_name, material_name, "Parametric", "Rectangle", f"{width * 1000:.3f}; {height * 1000:.3f}", "", "", area_m2, inertia, inertia, "", "", "", "", self._uuid(f"cross:{index}")])
            member_name = f"B{index}"
            member_rows.append([member_name, "Beam", cross_name, f"N{index}; N{index + 1}", "Line", float(span["length_m"]), "Line", "z by vector", 0.0, 0.0, 0.0, 1.0, "Centre", "Standard", "#FF2C8C99", self._uuid(f"member:{index}")])
            for case_id, value in span.get("load_case_values_kn_m", {}).items():
                load_rows.append([f"L_{index}_{case_id}", "Standard", "On beam", "Uniform", "Z", -abs(float(value)), member_name, str(case_id), "Global", "Length", "Relative", "From start", 0.0, 1.0, 0.0, 0.0, self._uuid(f"curve-load:{index}:{case_id}")])
        self._add_table(workbook, "StructuralCrossSection", ["Name", "Material", "Cross-section type", "Shape", "Parameters [mm]", "Profile", "Form code", "A [m2]", "Iy [m4]", "Iz [m4]", "It [m4]", "Iw[m6]", "Wply [m3]", "Wplz [m3]", "Id"], cross_rows)
        self._add_table(workbook, "StructuralCurveMember", ["Name", "Type", "Cross section", "Nodes", "Segments", "Length [m]", "Geometrical shape", "LCS", "LCS Rotation [deg]", "Coordinate X [m]", "Coordinate Y [m]", "Coordinate Z [m]", "System line", "Behaviour in analysis", "Color", "Id"], member_rows)
        support_rows = []
        for index, support in enumerate(model.get("supports", []), start=1):
            if not support.get("vertical") and not support.get("rotation"):
                continue
            support_rows.append([f"SUP{index}", "Fixed" if support.get("rotation") else "Hinged" if index == 1 else "Sliding", "In node", f"N{index}", "Rigid" if index == 1 else "Free", "Rigid", "Rigid" if support.get("vertical") else "Free", "Free", "Rigid" if support.get("rotation") else "Free", "Free", self._uuid(f"support:{index}")])
        self._add_table(workbook, "StructuralPointSupport", ["Name", "Type", "Boundary condition", "Node", "ux", "uy", "uz", "fix", "fiy", "fiz", "Id"], support_rows)
        self._add_table(workbook, "StructuralCurveAction", ["Name", "Type", "Force action", "Distribution", "Direction", "Value 1 [kN/m]", "Member", "Load case", "Coordinate system", "Location", "Coordinate definition", "Origin", "Start point [m]", "End point [m]", "Eccentricity ey [mm]", "Eccentricity ez [mm]", "Id"], load_rows)

    def _add_surface_model(self, workbook: Any, job: Mapping[str, Any], result: Mapping[str, Any] | None, material_name: str) -> None:
        model = job["analysis_model"]
        a = float(model["length_x_m"])
        b = float(model["length_y_m"])
        node_rows = [["N1", 0.0, 0.0, 0.0, self._uuid("surface-node:1")], ["N2", a, 0.0, 0.0, self._uuid("surface-node:2")], ["N3", a, b, 0.0, self._uuid("surface-node:3")], ["N4", 0.0, b, 0.0, self._uuid("surface-node:4")]]
        self._add_table(workbook, "StructuralPointConnection", ["Name", "Coordinate X [m]", "Coordinate Y [m]", "Coordinate Z [m]", "Id"], node_rows)
        self._add_table(workbook, "StructuralSurfaceMember", ["Name", "Type", "Material", "Thickness type", "Thickness [mm]", "System plane at", "Nodes", "Edges", "Area [m2]", "LCS Type", "Coordinate X [m]", "Coordinate Y [m]", "Coordinate Z [m]", "LCS Rotation [deg]", "Shape", "Behavior in analysis", "Color", "Id"], [["S1", "Plate", material_name, "Constant", float(model["thickness_m"]) * 1000.0, "Centre", "N1; N2; N3; N4", "Line; Line; Line; Line", a * b, "x by vector", 1.0, 0.0, 0.0, 0.0, "Flat", "Isotropic", "#7F2C8C99", self._uuid("surface:1")]])
        edge_rows = [[f"ES{edge}", "Hinged", "On edge", "S1", edge, "Free", "Free", "Rigid", "Free", "Free", "Free", "Global", "Relative", "From start", 0.0, 1.0, self._uuid(f"edge-support:{edge}")] for edge in range(1, 5)]
        self._add_table(workbook, "StructuralEdgeConnection", ["Name", "Type", "Boundary condition", "2D Member", "Edge", "ux", "uy", "uz", "fix", "fiy", "fiz", "Coordinate system", "Coordinate definition", "Origin", "Start point [m]", "End point [m]", "Id"], edge_rows)
        load_rows = [[f"SL_{case_id}", "Z", "Standard", "On 2D member", -abs(float(value)), "S1", str(case_id), "Global", "Length", self._uuid(f"surface-load:{case_id}")] for case_id, value in model.get("load_case_values_kn_m2", {}).items()]
        self._add_table(workbook, "StructuralSurfaceAction", ["Name", "Direction", "Type", "Force action", "Value [kN/m2]", "2D Member", "Load case", "Coordinate system", "Location", "Id"], load_rows)

    def _add_combinations(self, workbook: Any, result: Mapping[str, Any] | None) -> None:
        combinations = (((result or {}).get("load_combinations") or {}).get("combinations") or [])
        max_cases = max((len(item.get("factors", {})) for item in combinations), default=0)
        headers = ["Name", "Description", "Category", "Type"]
        for index in range(1, max_cases + 1):
            headers.extend([f"Load factor {index}", f"Multiplier {index}", f"Load case name {index}"])
        headers.append("Id")
        rows = []
        for combination in combinations:
            category = "ULS (Ultimate Limit State)" if combination["limit_state"] == "ULS" else "SLS (Serviceability Limit State)"
            row: list[Any] = [combination["combination_id"], combination["label"], category, "Linear"]
            for case_id, factor in combination.get("factors", {}).items():
                row.extend([float(factor), 1.0, case_id])
            while len(row) < len(headers) - 1:
                row.extend(["", "", ""])
            row.append(self._uuid(f"combination:{combination['combination_id']}"))
            rows.append(row)
        self._add_table(workbook, "StructuralLoadCombination", headers, rows)
