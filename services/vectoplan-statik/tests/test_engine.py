from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook
from pypdf import PdfReader

from src.design import PrestressTendonDesign, ReinforcedConcreteDesign, RetainingWallEarthPressureDesign
from src.design import MemberStabilityDesign
from src.domain import MaterialKind, StructureType
from src.exchange import Saf22Exporter, build_neutral_exchange
from src.loads import CombinationEngine, LoadCase, LoadPathBuilder, ThermalMovementCalculator
from src.materials import default_material_catalog
from src.pipeline import CalculationPipeline
from src.projects import ProjectCalculationPipeline, ProjectCaseRepository
from src.knowledge import LiteratureRegistry
from src.reference_cases import ReferenceCaseRepository
from src.reports.renderer import StructuralReportRenderer
from src.reports.dossier import StructuralDossierBuilder
from src.solvers import BeamLineSolver, GrillagePlateSolver, NavierPlateSolver, Truss2DSolver
from src.standards import default_standards_registry


def test_simply_supported_beam_matches_closed_form_solution():
    result = BeamLineSolver().solve(
        [{"span_id": "S1", "length_m": 5.0, "elastic_modulus_mpa": 210000.0, "inertia_m4": 8e-5, "uniform_load_kn_m": 10.0}]
    )
    assert result["nodes"][0]["vertical_reaction_kn"] == pytest.approx(25.0)
    assert result["nodes"][1]["vertical_reaction_kn"] == pytest.approx(25.0)
    assert result["envelope"]["max_abs_moment_knm"] == pytest.approx(10 * 5**2 / 8)
    expected_deflection_mm = 5 * 10 * 5**4 / (384 * 210000 * 1000 * 8e-5) * 1000
    assert result["envelope"]["max_abs_deflection_mm"] == pytest.approx(expected_deflection_mm, rel=1e-5)


def test_two_span_beam_is_in_equilibrium_and_has_hogging_support_moment():
    result = BeamLineSolver().solve(
        [
            {"span_id": "S1", "length_m": 6.0, "elastic_modulus_mpa": 210000.0, "inertia_m4": 0.001, "uniform_load_kn_m": 10.0},
            {"span_id": "S2", "length_m": 6.0, "elastic_modulus_mpa": 210000.0, "inertia_m4": 0.001, "uniform_load_kn_m": 10.0},
        ]
    )
    reactions = sum(node["vertical_reaction_kn"] for node in result["nodes"])
    assert reactions == pytest.approx(120.0)
    assert result["spans"][0]["end_actions"]["right_moment_knm"] < 0
    assert result["spans"][1]["end_actions"]["left_moment_knm"] < 0


def test_plate_field_is_symmetric_and_has_zero_supported_edges():
    result = NavierPlateSolver().solve(length_x_m=5.0, length_y_m=5.0, thickness_m=0.2, elastic_modulus_mpa=30000.0, poisson_ratio=0.2, uniform_load_kn_m2=10.0, terms=19, grid_size=21)
    rows = result["grid"]["rows"]
    center = rows[10][10]["w_mm"]
    assert center == pytest.approx(result["envelope"]["max_abs_deflection_mm"])
    assert rows[0][10]["w_mm"] == pytest.approx(0.0, abs=1e-9)
    assert rows[10][0]["w_mm"] == pytest.approx(0.0, abs=1e-9)
    assert rows[8][7]["w_mm"] == pytest.approx(rows[12][13]["w_mm"], rel=1e-7)


def test_en1990_oriented_combination_exposes_every_factor():
    profile = default_standards_registry().profile("DE_EC_2021")
    result = CombinationEngine(profile).generate([
        LoadCase("G", "Permanent", "permanent", 10.0, "kN/m"),
        LoadCase("Q", "Imposed", "variable", 5.0, "kN/m"),
    ])
    uls = next(item for item in result["combinations"] if item["combination_id"] == "ULS-001")
    assert uls["factors"] == {"G": 1.35, "Q": 1.5}
    assert uls["value"] == pytest.approx(21.0)


def test_concrete_standard_generation_is_selected_explicitly():
    registry = default_standards_registry()
    first_generation, _ = registry.select(
        profile_id="DE_EC_2021",
        structure_type=StructureType.RESIDENTIAL,
        material_kind=MaterialKind.REINFORCED_CONCRETE,
    )
    preview, _ = registry.select(
        profile_id="EU_2G_PREVIEW",
        structure_type=StructureType.RESIDENTIAL,
        material_kind=MaterialKind.REINFORCED_CONCRETE,
    )
    assert next(item for item in first_generation if item.role == "concrete_design").edition.startswith("2011")
    assert next(item for item in preview if item.role == "concrete_design").edition == "2025-09"


def test_concrete_design_applies_minimum_reinforcement_and_returns_trace():
    result = ReinforcedConcreteDesign(default_material_catalog()).check({
        "concrete_class": "C25/30", "reinforcement_class": "B500B", "width_mm": 1000.0,
        "height_mm": 200.0, "cover_mm": 30.0, "bar_diameter_mm": 10.0,
        "design_moment_knm": 8.0, "design_shear_kn": 20.0, "provided_reinforcement_mm2": 754.0,
    })
    assert result["required_reinforcement_mm2"] >= result["minimum_reinforcement_mm2"]
    assert {item["check_id"] for item in result["checks"]} == {"rc_bending", "rc_shear", "rc_ductility"}
    assert len(result["calculation_steps"]) == 3


def test_prestress_model_reduces_force_by_each_declared_loss():
    result = PrestressTendonDesign(default_material_catalog()).check({
        "prestress_grade": "Y1860S7", "area_mm2": 2850.0, "length_m": 42.0,
        "angular_change_rad": 0.22, "friction_mu": 0.19, "wobble_1_m": 0.006,
        "anchorage_slip_mm": 6.0, "relaxation_percent": 2.5,
    })
    forces = result["forces_kn"]
    assert forces["jack"] > forces["after_friction"] > forces["after_anchorage_slip"] > forces["effective_after_relaxation"] > 0


@pytest.mark.parametrize("case_id", ["single_span_rc", "two_span_steel", "residential_plate", "industrial_hall", "bridge_beam", "prestress_tendon", "masonry_wall", "pad_foundation", "bridge_abutment"])
def test_every_reference_case_runs_reproducibly(case_id):
    job = ReferenceCaseRepository().get(case_id)
    first = CalculationPipeline().run(job)
    second = CalculationPipeline().run(job)
    assert first["analysis_ref"] == second["analysis_ref"]
    assert first["summary"]["check_count"] >= 1
    assert first["verification"]["certified"] is False
    assert first["decisions"]
    assert first["calculation_steps"]


def test_html_and_pdf_reports_contain_traceable_results():
    repository = ReferenceCaseRepository()
    job = repository.get("residential_plate")
    result = CalculationPipeline().run(job)
    renderer = StructuralReportRenderer()
    html = renderer.render_html(job, result)
    assert "Warum wurde so gerechnet?" in html
    assert "/static/statik/css/report.css" in html
    assert "/static/statik/js/report.js" in html
    assert "<style>" not in html
    assert "Veränderbare Projektvariablen" in html
    pdf = renderer.render_pdf(job, result)
    assert pdf.startswith(b"%PDF")
    reader = PdfReader(BytesIO(pdf))
    assert len(reader.pages) >= 2
    assert "Normenbasis" in "".join(page.extract_text() or "" for page in reader.pages)


def test_dossier_contains_same_calculation_chain_used_by_reports():
    job = ReferenceCaseRepository().get("residential_plate")
    result = CalculationPipeline().run(job)
    dossier = StructuralDossierBuilder().build(job, result)
    assert dossier["contract_version"] == "structural-calculation-dossier/0.1"
    assert [item["check_id"] for item in dossier["checks"]] == [item["check_id"] for item in result["design"]["checks"]]
    assert all("comparison" in item for item in dossier["checks"])
    assert dossier["load_path"]["transfers"][0]["rule"] == "F = q · A"
    assert dossier["visualizations"][0]["kind"] == "surface_field"
    assert dossier["calculation_plan"]["formula_coverage"]["gate"]["passed"] is True
    assert {chapter["chapter_id"] for chapter in dossier["chapters"]} == {"01", "02", "03", "04", "05", "06", "07", "08", "09"}


def test_bridge_temperature_movement_is_explicit_and_checked():
    result = ThermalMovementCalculator().calculate({
        "length_m": 48.0, "thermal_expansion_1_k": 1e-5,
        "delta_t_positive_k": 30.0, "delta_t_negative_k": 20.0,
        "bearing_positive_capacity_mm": 20.0, "bearing_negative_capacity_mm": 15.0,
    })
    assert result["expansion_mm"] == pytest.approx(14.4)
    assert result["contraction_mm"] == pytest.approx(9.6)
    assert all(check["status"] == "passed" for check in result["checks"])


def test_retaining_wall_earth_and_water_resultants_are_auditable():
    result = RetainingWallEarthPressureDesign().check({
        "height_m": 6.0, "soil_unit_weight_kn_m3": 18.0, "friction_angle_deg": 30.0,
        "surcharge_kn_m2": 10.0, "water_depth_m": 2.0, "design_horizontal_resistance_kn_m": 300.0,
    })
    assert result["earth_pressure_coefficient_ka"] == pytest.approx(1 / 3, rel=1e-4)
    assert result["resultants"]["water_kn_m"] == pytest.approx(20.0)
    assert result["checks"][0]["design_value"] == result["resultants"]["total_horizontal_kn_m"]


def test_beam_load_path_reactions_close_governing_equilibrium():
    job = ReferenceCaseRepository().get("two_span_steel")
    result = CalculationPipeline().run(job)
    trace = LoadPathBuilder().build(job, result)
    reactions = [item["value"] for item in trace["transfers"] if item["status"] == "calculated_reaction"]
    governing = next(item for item in result["analysis"]["analyses"] if item["combination"]["limit_state"] == "ULS")
    expected = sum(span["uniform_load_kn_m"] * span["length_m"] for span in [
        {"uniform_load_kn_m": sum(float(value) * float(governing["combination"]["factors"].get(case, 0.0)) for case, value in source["load_case_values_kn_m"].items()), "length_m": source["length_m"]}
        for source in job["analysis_model"]["spans"]
    ])
    assert sum(reactions) == pytest.approx(expected, abs=1e-3)


def test_saf_export_uses_official_object_list_names():
    repository = ReferenceCaseRepository()
    job = repository.get("two_span_steel")
    result = CalculationPipeline().run(job)
    workbook = load_workbook(BytesIO(Saf22Exporter().to_bytes(job, result)))
    expected = {"Project", "Model", "StructuralMaterial", "StructuralPointConnection", "StructuralCurveMember", "StructuralPointSupport", "StructuralCurveAction", "StructuralLoadCase", "StructuralLoadCombination"}
    assert expected.issubset(workbook.sheetnames)
    assert "StructuralCurveMember" in workbook["StructuralCurveMember"].tables
    assert workbook["Model"]["B13"].value == "2.2.0"


def test_neutral_exchange_declares_ifc_limit_and_saf_status():
    job = ReferenceCaseRepository().get("single_span_rc")
    result = CalculationPipeline().run(job)
    exchange = build_neutral_exchange(job, result)
    assert exchange["adapter_targets"]["SAF"]["version"] == "2.2.0"
    assert exchange["adapter_targets"]["SAF"]["round_trip_certified"] is False
    assert exchange["adapter_targets"]["IFC"]["implemented"] is False


def test_grillage_handles_opening_and_internal_supports():
    result = GrillagePlateSolver().solve(
        length_x_m=8.0, length_y_m=6.0, thickness_m=0.22,
        elastic_modulus_x_mpa=33000, elastic_modulus_y_mpa=28000,
        uniform_load_kn_m2=9.0, nx=9, ny=7,
        openings=[{"x_min_m": 3.0, "x_max_m": 5.0, "y_min_m": 2.0, "y_max_m": 4.0}],
        line_supports=[{"axis": "x", "coordinate_m": 3.0, "start_m": 0.0, "end_m": 8.0}],
        point_supports=[{"x_m": 6.0, "y_m": 2.0}],
        cracking={"cracking_moment_knm_m": 3.0, "cracked_stiffness_factor": 0.5},
    )
    assert abs(result["equilibrium_residual_kn"]) < 1e-4
    assert any(cell["active"] is False for row in result["grid"]["rows"] for cell in row)
    assert result["iterations"][-1]["stiffness_factor"] < 1.0
    assert {"openings", "line_supports", "point_supports", "orthotropic", "secant_cracking"}.issubset(result["applicability"]["supported"])


def test_roof_truss_is_in_equilibrium():
    project = ProjectCalculationPipeline().run(ProjectCaseRepository().get("complex_residential_building"))
    roof = next(item for item in project["positions"] if item["position_ref"] == "P07")
    analysis = roof["result"]["analysis"]["analyses"][0]["result"]
    assert sum(item["ry_kn"] for item in analysis["nodes"]) == pytest.approx(86.4, abs=1e-5)
    assert analysis["members"]


def test_foundation_contact_and_member_stability_are_explicit():
    result = MemberStabilityDesign().check({
        "length_m": 3.0, "effective_length_factor": 1.0, "elastic_modulus_mpa": 210000,
        "inertia_m4": 8e-5, "design_axial_kn": 300, "first_order_moment_knm": 12,
        "design_moment_resistance_knm": 80,
    })
    assert result["critical_load_kn"] > 300
    assert result["magnification_factor"] > 1
    assert {item["check_id"] for item in result["checks"]} == {"elastic_euler_stability", "second_order_moment"}


def test_complex_residential_project_runs_every_position():
    result = ProjectCalculationPipeline().run(ProjectCaseRepository().get("complex_residential_building"))
    assert result["summary"]["position_count"] == 9
    assert result["summary"]["check_count"] >= 20
    assert result["summary"]["unresolved_capability_count"] == 0
    assert result["summary"]["status"] == "passed"
    assert {item["position_ref"] for item in result["positions"]} == {f"P{index:02}" for index in range(1, 10)}
    assert all(item["result"]["calculation_steps"] for item in result["positions"])


def test_literature_catalog_links_every_topic_to_backend_and_tests():
    records = LiteratureRegistry().records()
    assert len(records) == 12
    assert all(item["backend"] and item["tests"] and item["book_pages"] for item in records)
    assert {item["topic_id"] for item in records} >= {"equilibrium_loads", "buckling", "trusses", "worked_house"}


def test_new_api_catalogs_jobs_reports_and_exchange(client):
    assert client.get("/api/v1/statik/standards").status_code == 200
    assert client.get("/api/v1/statik/materials").status_code == 200
    assert len(client.get("/api/v1/statik/reference-cases").get_json()["cases"]) == 9
    run = client.get("/api/v1/statik/reference-cases/single_span_rc/run")
    assert run.status_code == 200
    assert run.get_json()["contract_version"] == "structural-analysis-result/0.2"
    html = client.get("/api/v1/statik/reference-cases/single_span_rc/report.html")
    assert html.status_code == 200
    assert html.mimetype == "text/html"
    pdf = client.get("/api/v1/statik/reference-cases/single_span_rc/report.pdf")
    assert pdf.status_code == 200
    assert pdf.data.startswith(b"%PDF")
    saf = client.get("/api/v1/statik/reference-cases/single_span_rc/exchange.saf")
    assert saf.status_code == 200
    assert saf.data[:2] == b"PK"
    dossier = client.get("/api/v1/statik/reference-cases/residential_plate/dossier.json")
    assert dossier.status_code == 200
    assert dossier.get_json()["contract_version"] == "structural-calculation-dossier/0.1"
